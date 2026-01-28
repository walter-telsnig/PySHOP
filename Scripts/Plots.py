from Misc import *
import matplotlib
from openpyxl import load_workbook

# setting date format for xaxis labels
formatter = matplotlib.dates.DateFormatter('%d/%m\n%Y')

res_plot_scaling = 0.05

def plot_storage_results(shop, starttime=pd.Timestamp(2023, 1, 1), endtime=pd.Timestamp(2024, 1, 1), hide_plots=True):

    if hide_plots:
        plt.ioff()

    d = shop.get_time_resolution()
    shop.set_time_resolution(starttime=d['starttime'], endtime=d['endtime'], timeunit=d['timeunit'],
                             timeresolution=pd.Series(index=[d['starttime']], data=[d['timeresolution'].values[0]]))
    d = shop.get_time_resolution()
    t = d['timeresolution'].index[0:-1]

    list_of_objects = shop.model.reservoir.get_object_names()

    rows, cols = return_number_of_rows_and_cols(list_of_objects)

    fig, ax_arr = plt.subplots(rows, cols)
    fig.set_figwidth(18)
    fig.set_figheight(8)

    if isinstance(ax_arr,np.ndarray):
        ax_arr = ax_arr.flatten()
    else:
        ax_arr = np.array([ax_arr])

    wb = load_workbook(cubature_file, read_only=True) #loading cubature files to check later if cubature for a given
                                                      # reservoir exists and limits can be converted

    counter = 0
    for ax in ax_arr:
        if counter < len(list_of_objects):
            res = shop.model.reservoir[list_of_objects[counter]]
            ax.set_title(list_of_objects[counter])
            ax.fill_between(t, res.head.get().values[:-1], 0, color='blue', label='Storage')

            if res.get_name() in wb.sheetnames:
                if not res.tactical_limit_min.get() is None:
                    ax.plot(t, convert_vol_to_head(res.tactical_limit_min.get().values, res.get_name()), color='orange')
                if not res.tactical_limit_max.get() is None:
                    ax.plot(t, convert_vol_to_head(res.tactical_limit_max.get().values, res.get_name()), color='orange')
                if not res.min_head_constr.get() is None:
                    ax.plot(t, res.min_head_constr.get().values, color='red', label='min head constr.')
                if not res.max_head_constr.get() is None:
                    ax.plot(t, res.max_head_constr.get().values, color='red', label='max head constr.')
            ax.set_ylabel(r'Level a.A. [m]')

            lrl = res.lrl.get()
            hrl = res.hrl.get()
            dif_res_level = hrl-lrl

            ax.set_ylim(lrl - dif_res_level * res_plot_scaling, res.hrl.get() + dif_res_level * res_plot_scaling)
            ax.set_xlim(starttime, endtime)
            ax.xaxis.set_major_formatter(formatter)

            if not res.inflow.get() is None and len(res.inflow.get().values) == len(t):
                ax_twin = ax.twinx()
                ax_twin.plot(t, res.inflow.get().values, color='black', label='Inflow')
                ax_twin.set_ylabel(r'Inflow [m$^3$/s]')
        else:
            ax.axis('off')

        counter += 1

    plt.tight_layout()

    if hide_plots:
        plt.ion()



def plot_production_results(shop, starttime=pd.Timestamp(2023, 1, 1), endtime=pd.Timestamp(2024, 1, 1), daily_avg_price=True,hide_plots=True):

    if hide_plots:
        plt.ioff()

    d = shop.get_time_resolution()
    shop.set_time_resolution(starttime=d['starttime'], endtime=d['endtime'], timeunit=d['timeunit'],
                             timeresolution=pd.Series(index=[d['starttime']], data=[d['timeresolution'].values[0]]))
    d = shop.get_time_resolution()
    t = d['timeresolution'].index[0:-1]

    if daily_avg_price:
        plot_prices, dates = get_daily_avg_prices(shop)
    else:
        market_name = shop.model.market.get_object_names()[0]
        plot_prices = shop.model.market[market_name].sale_price.get().values
        dates = t

    list_of_objects = shop.model.plant.get_object_names()

    rows, cols = return_number_of_rows_and_cols(list_of_objects)

    fig, ax_arr = plt.subplots(rows, cols)
    fig.set_figwidth(18)
    fig.set_figheight(8)

    counter = 0

    if isinstance(ax_arr,np.ndarray):
        ax_arr = ax_arr.flatten()
    else:
        ax_arr = np.array([ax_arr])

    for ax in ax_arr:
        if counter < len(list_of_objects):
            res = shop.model.plant[list_of_objects[counter]]
            ax.set_title("KW " + list_of_objects[counter])
            ax.fill_between(t, res.production.get().values, color='red', step='mid', label='Production')
            ax.fill_between(t, -res.consumption.get().values, color='blue', step='mid', label='Consumption')
            ax.set_ylabel(r'Prod./Cons. [MW]')
            ax.set_xlim(starttime, endtime)
            ax.xaxis.set_major_formatter(formatter)

            ax_twin = ax.twinx()
            ax_twin.plot(dates, plot_prices, color='black')
            ax_twin.set_ylabel(r'Price [€]')
        else:
            ax.axis('off')

        counter += 1

    plt.tight_layout()

    if hide_plots:
        plt.ion()