import pandas as pd
import numpy as np
import pathlib
import matplotlib.pyplot as plt
import time
import os
from openpyxl import load_workbook

cubature_file = "C:/pySHOP/Input/Cubatures/Powel.xlsx"


########################################################################################################################
# ! Set global penalty value for violating tactical min/max
########################################################################################################################
def get_limit_val():
    return 1e4


def create_date_string(timestamp):
    y = str(timestamp.year)
    m = str(timestamp.month)
    d = str(timestamp.day)

    if len(m) != 2:
        m = "0" + m

    if len(d) != 2:
        d = "0" + d

    return y + m + d

########################################################################################################################
# Total volume check
# Measures the total water volume at the beginning and the end of the simulation and compares its difference to the
# total inflow during the simulation
########################################################################################################################
def total_volume_check(shop,save_path="",detailed=False):
    res_list = shop.model.reservoir.get_object_names()

    timeunit = shop.get_time_resolution()["timeunit"]
    timeresolution = shop.get_time_resolution()["timeresolution"].values[0]

    #scaling factor to convert m3/s inflow data to total inflow volume per timestep in the simulation
    scaling_f = 1
    if timeunit == "minute":
        scaling_f = timeresolution * 60
    elif timeunit == "hour":
        scaling_f = timeresolution * 3600
    else:
        print("Timeunit not defined.")

    start_vol = 0
    end_vol = 0
    inflow = 0

    for res in res_list:
        start_vol += shop.model.reservoir[res].storage.get().values[0]
        end_vol += shop.model.reservoir[res].storage.get().values[-1]

        if not res in ["Drau","Moell","Woerthersee"]:
            inflow += np.sum(shop.model.reservoir[res].inflow.get().values)/1e6*scaling_f

    if detailed:
        print("############################################################")
        print("# Total volume check")
        print("############################################################")
        print("Start: ".ljust(30)                   + ("{:.3f}".format(start_vol) + " Mm3").rjust(30))
        print("End: ".ljust(30)                     + ("{:.3f}".format(end_vol) + " Mm3").rjust(30))
        print("Volume Change: ".ljust(30)           + ("{:.3f}".format((end_vol - start_vol)) + " Mm3").rjust(30))
        print("Inflow: ".ljust(30)                  + ("{:.3f}".format(inflow) + " Mm3").rjust(30))
        print("Volume Change - Inflow: ".ljust(30)  + ("{:.2e}".format((end_vol - start_vol)-inflow) + " Mm3").rjust(30))
    else:
        print("############################################################")
        print("# Total volume check")
        print("Volume Change - Inflow: ".ljust(30)  + ("{:.2e}".format((end_vol - start_vol)-inflow) + " Mm3").rjust(30))

    if save_path != "":
        f = open(save_path+"volume_check.txt", "w")
        f.write("Start: ".ljust(30) + ("{:.3f}".format(start_vol) + " Mm3").rjust(15)+"\n")
        f.write("End: ".ljust(30) + ("{:.3f}".format(end_vol) + " Mm3").rjust(15)+"\n")
        f.write("Volume Change: ".ljust(30) + ("{:.3f}".format((end_vol - start_vol)) + " Mm3").rjust(15)+"\n")
        f.write("Inflow: ".ljust(30) + ("{:.3f}".format(inflow) + " Mm3").rjust(15)+"\n")
        f.write("Volume Change - Inflow: ".ljust(30) + ("{:.2e}".format((end_vol - start_vol) - inflow) + " Mm3").rjust(15)+"\n")
        f.close()

########################################################################################################################
# Returns a time series of the daily averaged price for any time resolution
########################################################################################################################
def get_daily_avg_prices(shop):
    timeSeries = shop.get_time_resolution()['timeresolution']

    starttime = timeSeries.index[0]
    endtime   = timeSeries.index[-1]

    timeunit = shop.get_time_resolution()["timeunit"]
    timeresolution = shop.get_time_resolution()["timeresolution"].values[0]
    n_days = int(np.round((endtime - starttime) / np.timedelta64(1, 'D')))

    # calculate the number of timesteps within a day given the simulations time
    if timeunit == "hour":
        steps = int(24 / timeresolution)
    elif timeunit == "minute":
        steps = int(24 * 60 / timeresolution)
    else:
        print("Timeunit not defined.")

    market_name = shop.model.market.get_object_names()[0]

    plot_prices = [np.mean(shop.model.market[market_name].sale_price.get().values[x * steps:(x + 1) * steps]) for x in
                   range(n_days)]
    dates = [starttime + pd.Timedelta(x, unit='D') for x in range(n_days)]

    return plot_prices, dates


########################################################################################################################
# Show the runtime of a script
# Prints the time relative to a given start time (run_start)
# If a save path is provided, the runtime is written (or appended if it already exists) to a text file
########################################################################################################################
def show_runtime(run_start,save_path=""):

    runtime = (time.time() - run_start)/60
    print("############################################################")
    print("# Runtime: " + "{:.2f}".format(runtime) + " min")
    print("############################################################")

    if save_path != "":
        f = open(save_path+"runtime.txt", "a")
        f.write("Runtime: " + "{:.2f}".format(runtime) + " min\n")
        f.close()

########################################################################################################################
# Writes the time series to an excel file
# Reservoirs:   - head in [m]asl
#               - storage in [Mm3]
#               - tac. min and max in [Mm3] and [m]asl
#               - min and max head constraint in [m]asl
# Plants:       - production [MW]
#               - consumption [MW]
#               - discharge [m3/s /h]
########################################################################################################################
def create_dataframe_for_export(shop,save_path,topology):
    timeSeries = shop.get_time_resolution()['timeresolution']

    market_name = shop.model.market.get_object_names()[0]
    price = shop.model.market[market_name].sale_price.get().values

    results = pd.DataFrame({"Price [€]": price},index=timeSeries[:-1].index).copy()
    # "copy" to avert the following warning
    # PerformanceWarning: DataFrame is highly fragmented.  This is usually the result of calling `frame.insert` many
    # times, which has poor performance.  Consider joining all columns at once using pd.concat(axis=1) instead.
    # To get a de-fragmented frame, use `newframe = frame.copy()`

    list_of_reservoirs  = shop.model.reservoir.get_object_names()
    list_of_powerplants = shop.model.plant.get_object_names()
    list_of_gates = shop.model.gate.get_object_names()

    wb = load_workbook(cubature_file, read_only=True) #loading cubature files to check later if cubature for a given
                                                      # reservoir exists and limits can be converted

    for plant in list_of_powerplants:
        results["KW_" + plant + " Production [MW]"] = shop.model.plant[plant].production.get().values
        results["KW_" + plant + " Consumption [MW]"] = shop.model.plant[plant].consumption.get().values
        results["KW_" + plant + " Production value [€]"] = shop.model.plant[plant].production.get().values * price
        results["KW_" + plant + " Consumption value [€]"] = shop.model.plant[plant].consumption.get().values * price
        results["KW_" + plant + " Discharge [m3/s]"] = shop.model.plant[plant].discharge.get().values
        results["KW_" + plant + " Upflow [m3/s]"] = shop.model.plant[plant].upflow.get().values

    for res in list_of_reservoirs:
        results[res + " level a.A. [m]"] = shop.model.reservoir[res].head.get().values[:-1]
        results[res + " storage [Mm3]"]  = shop.model.reservoir[res].storage.get().values[:-1]
        if not res in ["Moell", "Drau", "Woerthersee"]:
            if res in wb.sheetnames and not shop.model.reservoir[res].tactical_limit_min.get() is None:
                results[res + " tac. min [Mm3]"] = shop.model.reservoir[res].tactical_limit_min.get().values
                results[res + " tac. min [m]"] = convert_vol_to_head(shop.model.reservoir[res].tactical_limit_min.get().values, res)
            if res in wb.sheetnames and not shop.model.reservoir[res].tactical_limit_max.get() is None:
                results[res + " tac. max [Mm3]"] = shop.model.reservoir[res].tactical_limit_max.get().values
                results[res + " tac. max [m]"] = convert_vol_to_head(shop.model.reservoir[res].tactical_limit_max.get().values, res)
            if not shop.model.reservoir[res].min_head_constr.get() is None:
                results[res + " min head constr. [m]"] = shop.model.reservoir[res].min_head_constr.get().values
            if not shop.model.reservoir[res].max_head_constr.get() is None:
                results[res + " max head constr. [m]"] = shop.model.reservoir[res].max_head_constr.get().values
            if not shop.model.reservoir[res].inflow.get() is None:
                results[res + " inflow [m3/s]"] = shop.model.reservoir[res].inflow.get().values
            if res in wb.sheetnames and not shop.model.reservoir[res].schedule.get() is None:
                results[res + " schedule [Mm3]"] = shop.model.reservoir[res].schedule.get().values[:-1]

    for gate in list_of_gates:
        results[gate + " discharge [m3/s]"] = shop.model.gate[gate].discharge.get().values

    sd = shop.get_time_resolution()['timeresolution'].index[0]
    ds = create_date_string(sd)
    results.to_excel(save_path + "timeSeries_"+topology+"_"+ds+".xlsx")

########################################################################################################################
# Load hpfc
########################################################################################################################
def load_hpfc(file):

    df = pd.read_excel("C:/pySHOP/Input/Price_Curve/"+file, sheet_name="AT")
    df.drop_duplicates(inplace=True)
    prices = df["Price"].tolist()
    deliv_date = df["Delivery Date"].tolist()

    return deliv_date, prices

########################################################################################################################
# Load market data
########################################################################################################################
def load_market_data(shop,file="HPFCs Budget 2023"):
#def load_market_data(shop,file="HPFC_AT_2024_05_09_2024_2025"):

    starttime = shop.get_time_resolution()['timeresolution'].index[0]

    shop.model.market.add_object('price_curve')
    price_curve = shop.model.market.price_curve
    dates, prices = load_hpfc(file)
    price_curve.sale_price.set(pd.DataFrame(prices, index=dates))

    price_curve.buy_price.set(price_curve.sale_price.get())
    price_curve.max_buy.set(pd.Series([9999], [starttime]))
    price_curve.max_sale.set(pd.Series([9999], [starttime]))
    price_curve.load.set(pd.Series([0], [starttime]))

    return shop

########################################################################################################################
# run simulation
########################################################################################################################
def run_opt(shop,run_start=time.time(),save_path="",code1 = "head", n_iterations1 = "5", code2="incremental",n_iterations2="3",check_volume=True):
    shop.set_code([code1], [])
    shop.start_sim([], [n_iterations1])
    shop.set_code([code2], [])
    shop.start_sim([], [n_iterations2])

    if check_volume:
        total_volume_check(shop, save_path=save_path)
    show_runtime(run_start, save_path=save_path)

########################################################################################################################
# get a time series that changes twice: 1h after the start and 1h before the end
# was used to set limits at the start/end so storage ends at a specific level
########################################################################################################################
def get_limit_time_series(starttime,endtime):
    limit_delta = 1
    limit_time_series = [starttime, starttime + pd.Timedelta(limit_delta, unit='H'),
                         endtime - pd.Timedelta(limit_delta, unit='H')]
    return limit_time_series

########################################################################################################################
# load cubature
########################################################################################################################
def detail_kubatur(sheet):

    df = pd.read_excel(cubature_file, sheet_name=sheet)

    # if using Kubaturen.xlsx -> more detailed cubatures
    #storage = df["V [m3]"].values / 1e6
    #level = df["masl [m]"].values

    storage = df["storage"].values
    level = df["water level"].values

    return storage, level

########################################################################################################################
# get the volume corresponding to a [m]asl value for a given cubature
########################################################################################################################
def kubatur(value, sheet):

    pathlib.Path(cubature_file).exists()
    df = pd.read_excel(file, sheet_name=sheet)

    mask = np.where(df['masl [m]'] == value)
    print(df.iloc[mask])

########################################################################################################################
# transforms tactical limits from volume [Mm3] to [m]asl
########################################################################################################################
def convert_vol_to_head(series, res):

    storage, level = detail_kubatur(res)

    head_level = np.zeros(len(series))

    for val in np.arange(len(series)):
        # linear interpolation
        x = series[val]
        x1 = storage[np.where(series[val] > storage)[0][-1]]
        x2 = storage[np.where(series[val] <= storage)[0][0]]
        y1 = level[np.where(series[val] > storage)[0][-1]]
        y2 = level[np.where(series[val] <= storage)[0][0]]
        head_level[val] = y1 + (x-x1)*(y2-y1)/(x2-x1)

    return head_level

def convert_head_to_vol(res, head):
    storage, level = detail_kubatur(res)

    # linear interpolation
    x = head
    x1 = level[np.where(level < head)[0][-1]]
    x2 = level[np.where(level >= head)[0][0]]
    y1 = storage[np.where(level < head)[0][-1]]
    y2 = storage[np.where(level >= head)[0][0]]
    vol = y1 + (x - x1) * (y2 - y1) / (x2 - x1)

    return vol

########################################################################################################################
# Calculates the cubature and its first and second derivative to check its slope and convexity
########################################################################################################################
def convexity_check(res):
    storage, level = detail_kubatur(res)

    #y = keep_convexity(storage)
    #x = keep_convexity(level)

    y = storage
    x = level

    first_deriv = get_deriv(x,y)
    second_deriv = get_deriv(x,first_deriv)
    second_deriv[0]=np.nan
    second_deriv[-1]=np.nan

    n_neg_values = np.sum(np.where(second_deriv < 0,1,0))
    n_pos_values = np.sum(np.where(second_deriv > 0,1,0))
    print("Neg values: " + str(n_neg_values))
    print("Pos values: " + str(n_pos_values))

    if n_neg_values > 0:
        loc_neg_values = np.where(np.where(second_deriv < 0))
        print(loc_neg_values)

    fig = plt.figure()
    ax1 = plt.subplot(3,1,1)
    ax1.axhline(0,color='black',linestyle='--',alpha=0.5)
    ax1.plot(x,y)
    ax1.set_xlabel("x")
    ax1.set_ylabel("y(x)")
    ax1.set_title(res)


    ax2 = plt.subplot(3,1,2)
    ax2.axhline(0,color='black',linestyle='--',alpha=0.5)
    ax2.plot(x,first_deriv)
    ax2.set_xlabel("x")
    ax2.set_ylabel("y'(x)")

    ax3 = plt.subplot(3,1,3)
    ax3.axhline(0,color='black',linestyle='--',alpha=0.5)
    ax3.plot(x,second_deriv)
    ax3.set_xlabel("x")
    ax3.set_ylabel("y''(x)")

    plt.tight_layout()


########################################################################################################################
# very simple central derivative of a given timeseries in x and y with forward and backwards derivations at the
# beginning and end, respectively
########################################################################################################################
def get_deriv(x,y):
    length = len(x)

    deriv = np.zeros(length)

    deriv[0] = (y[1] - y[0]) / (x[1] - x[0])
    deriv[-1] = (y[-1] - y[-2]) / (x[-1] - x[-2])

    for i in np.arange(1, length-1):
        deriv[i] = (y[i+1]-y[i-1])/(x[i+1]-x[i-1])

    return deriv


########################################################################################################################
# Plotting the Powel (SIM) vs the more detailed cubatures
########################################################################################################################
def compare_cubature(sheet):
    file = "C:/pySHOP/Input/Cubatures/Powel.xlsx"
    df = pd.read_excel(cubature_file, sheet_name=sheet)

    storage_Powel = df["storage"].values
    level_Powel = df["water level"].values

    file = "C:/pySHOP/Input/Cubatures/Kubaturen.xlsx"
    df = pd.read_excel(file, sheet_name=sheet)

    storage_detail = df["V [m3]"].values / 1e6
    level_detail = df["masl [m]"].values

    plt.plot(level_Powel, storage_Powel, color='blue', label="Kubatur in Powel")
    plt.plot(level_detail, storage_detail, color='red', label="Kubatur detailliert")
    plt.axhline(0.025, color='black', linestyle='--', label='tact. limit max')
    plt.legend()


########################################################################################################################
# create path
########################################################################################################################
def set_path(name):
    file_path_root = "C:/pySHOP/Results/"+name+"/"
    file_path_detail = file_path_root + "Detail/"

    if not os.path.isdir(file_path_root):
        os.makedirs(file_path_root)
        os.makedirs(file_path_detail)
    elif not os.path.isdir(file_path_detail):
        os.makedirs(file_path_detail)

    return file_path_root, file_path_detail

########################################################################################################################
# CHECK FOR HYDRAULIC SHORT CIRCUIT, I.E. SIMULTANEOUS PRODUCTION AND CONSUMPTION
########################################################################################################################
def check_for_hydraulic_short_circuit(KW):

    prod = KW.production.get().values
    cons = KW.consumption.get().values

    HSC = np.where(np.logical_and(prod != 0, cons != 0))
    n_occurances = len(HSC[0])

    if n_occurances == 0:
        print("No hydraulic short circuit occurs.")
    else:
        print("Hydraulic short circuit occurs:")
        print(HSC)

########################################################################################################################
# create a time series spanning start and endtime in order to set the min. head. restriction to a certain level at the
# end of the year
########################################################################################################################
def year_end_head_limit(shop, base,top, end,use=True, spread = 0.5):

    starttime = shop.get_time_resolution()['timeresolution'].index[0]
    endtime = shop.get_time_resolution()['timeresolution'].index[-1]

    n_years = endtime.year - starttime.year + 1
    time_series = []
    min_limit = []
    max_limit = []

    offset = shop.get_time_resolution()['timeresolution'].values[0]

    for i in np.arange(0, n_years):
        y = starttime.year
        time_series.append(pd.Timestamp(y + i, 1, 1))
        min_limit.append(base)
        max_limit.append(top)
        if use:
            time_series.append(pd.Timestamp(y + (i + 1), 1, 1) - pd.Timedelta(offset, unit="H"))
            min_limit.append(end-spread)
            max_limit.append(end+spread)

    return time_series, min_limit, max_limit

def reoccuring_limit(shop, dates,values,res=""):

    starttime = shop.get_time_resolution()['timeresolution'].index[0]
    endtime = shop.get_time_resolution()['timeresolution'].index[-1]

    n_years = endtime.year - starttime.year + 1
    time_series = []
    limit = []

    y = starttime.year
    for i in np.arange(0, n_years):
        for j in np.arange(0,len(dates)):
            time_series.append(pd.Timestamp(y + i, dates[j][0], dates[j][1]))
            if res == "":
                limit.append(values[j])
            else:
                limit.append(convert_head_to_vol(res, values[j]))

    return time_series, limit

def get_tactical_limit(starttime,endtime,res):

    n_years = endtime.year - starttime.year + 1
    time_series = []
    min_limit = []

    for i in np.arange(0, n_years):
        y = starttime.year
        time_series.append(pd.Timestamp(y + i, 1, 1))
        time_series.append(pd.Timestamp(y + i, 6, 15))
        time_series.append(pd.Timestamp(y + i, 11, 1))
        if res == "Soboth":
            min_limit.append(convert_head_to_vol("Soboth",1060))
            min_limit.append(convert_head_to_vol("Soboth",1079))
            min_limit.append(convert_head_to_vol("Soboth",1060))
        elif res == "Freibach":
            min_limit.append(convert_head_to_vol("Freibach",706))
            min_limit.append(convert_head_to_vol("Freibach",720.5))
            min_limit.append(convert_head_to_vol("Freibach",706))
        elif res == "Forstsee":
            min_limit.append(convert_head_to_vol("Forstsee",588.63))
            min_limit.append(convert_head_to_vol("Forstsee",602.5))
            min_limit.append(convert_head_to_vol("Forstsee",588.63))
        else:
            print("Reservoir not defined.")

    return time_series, min_limit

def get_factor(KW):
    if KW == "Freibach":
        return 0.68
    elif KW == "Forstsee":
        return 0.58
    else:
        return 0.75

def get_economics(shop,save_path="",verbose=True):

    starttime = shop.get_time_resolution()['timeresolution'].index[0]
    endtime = shop.get_time_resolution()['timeresolution'].index[-1]

    price = shop.model.market.price_curve.sale_price.get().values

    list_of_plants = shop.model.plant.get_object_names()

    total_earnings = 0

    if save_path != "":
        f = open(save_path+"earnings.txt","w")
        f.write("############################################################"+"\n")
        f.write("# Economics overview"+"\n")
        f.write("# Start date:".ljust(20) + str(starttime).rjust(40)+"\n")
        f.write("# End date:".ljust(20) + str(endtime).rjust(40)+"\n")
        f.write("############################################################"+"\n")

    if verbose:
        print("############################################################")
        print("# Economics overview")
        print("# Start date:".ljust(20) + str(starttime).rjust(40))
        print("# End date:".ljust(20) + str(endtime).rjust(40))
        print("############################################################")

    for item in list_of_plants:
        KW = shop.model.plant[item]
        prod = KW.production.get().values
        cons = KW.consumption.get().values

        prod_euro = np.sum(prod * price)
        cons_euro = np.sum(cons * price)

        earnings = prod_euro - cons_euro
        total_earnings += earnings

        if save_path != "":
            f.write("# " + KW.get_name()+"\n")
            f.write("#-----------------------------------------------------------"+"\n")
            f.write("# Production: ".ljust(20) + (str(int(np.sum(prod))) + " MW").rjust(20) + (str(int(prod_euro)) + " €").rjust(20)+"\n")
            f.write("# Consumption: ".ljust(20) + (str(int(np.sum(cons))) + " MW").rjust(20) + (str(int(cons_euro)) + " €").rjust(20)+"\n")
            f.write("# Earnings: ".ljust(20) + (str(int(np.sum(prod)-np.sum(cons))) + " MW").rjust(20) + (str(int(earnings)) + " €").rjust(20)+"\n")
            f.write("############################################################"+"\n")

        if verbose:
            print("# " + KW.get_name())
            print("#-----------------------------------------------------------")
            print("# Production: ".ljust(20) + (str(int(np.sum(prod))) + " MW").rjust(20) + (str(int(prod_euro)) + " €").rjust(20))
            print("# Consumption: ".ljust(20) + (str(int(np.sum(cons))) + " MW").rjust(20) + (str(int(cons_euro)) + " €").rjust(20))
            print("# Earnings: ".ljust(20) + (str(int(earnings)) + " €").rjust(40))
            print("############################################################")

    if save_path != "":
        f.write("# Total earnings: ".ljust(20) + (str(int(total_earnings))+ " €").rjust(40)+"\n")
        f.write("############################################################"+"\n")
        f.close()
    if verbose:
        print("# Total earnings: ".ljust(20) + (str(int(total_earnings))+ " €").rjust(40))
        print("############################################################")



def convert_int_to_month(i):
    month_dict = {1: "Jan",
                  2: "Feb",
                  3: "Mar",
                  4: "Apr",
                  5: "May",
                  6: "Jun",
                  7: "Jul",
                  8: "Aug",
                  9: "Sep",
                  10: "Oct",
                  11: "Nov",
                  12: "Dec"}

    return month_dict[i]

def create_dir_name(starttime,endtime):

    start_y = str(starttime.year)
    start_m = convert_int_to_month(starttime.month)
    start_d = str(starttime.day)

    end_y = str(endtime.year)
    end_m = convert_int_to_month(endtime.month)
    end_d = str(endtime.day)

    return start_y + "-" + start_m + "-" + start_d + "_" + end_y + "-" + end_m + "-" + end_d


def return_number_of_rows_and_cols(list_of_objects):
    n = len(list_of_objects)

    side_length = 1
    while side_length * side_length < n:
        side_length += 1
        if side_length >= 1000:
            print("Error in return_number_of_rows_and_cols(): side_length exceeding 1000.")
            break

    if side_length * (side_length-1) >= n:
        cols = side_length-1
    else:
        cols = side_length

    rows = side_length

    return rows, cols


def read_command_file(shop, file):

    with open(file, "r") as file:
        commands = file.readlines()

    for line in commands:
        shop = execute_command(shop,line)

    return shop

def execute_command(shop,line):
    line = line.replace("\n", "")
    line = line.split(" ")
    line = [ele for ele in line if ele != ""]

    #currently not used: n_it_1, n_it_2
    n_it_1 = 0
    n_it_2 = 0

    if "time_delay_unit" in line:
        shop.model.global_settings.global_settings.time_delay_unit.set(line[2])
    elif "method" in line:
        shop.model.global_settings.global_settings.solver_algorithm.set(line[2][1:])
    elif "bypass_loss" in line:
        if line[2] == "/on":
            val = 1
        else:
            val = 0
        shop.model.global_settings.global_settings.bypass_loss.set(val)
    elif "penalty" in line:
        if "/endpoint" in line:
            shop.model.global_settings.global_settings.rsv_penalty_cost.set(eval(line[2]))
            shop.model.global_settings.global_settings.rsv_hard_limit_penalty_cost.set(eval(line[2]))
        elif "/ramping" in line:
            shop.model.global_settings.global_settings.production_ramp_penalty_cost.set(eval(line[2]))
        elif "/load" in line:
            shop.model.global_settings.global_settings.load_penalty_cost.set(eval(line[2]))
        elif "/discharge" in line:
            shop.model.global_settings.global_settings.discharge_group_penalty_cost.set(eval(line[2]))
        elif "/overflow" in line:
            shop.model.global_settings.global_settings.overflow_cost.set(eval(line[2]))

    elif "mipgap" in line:
        shop.model.global_settings.global_settings.mipgap_rel.set(eval(line[2]))
    elif "timelimit" in line:
        shop.model.global_settings.global_settings.timelimit.set(eval(line[2]))
    elif "sim" in line:
        count = int(line[2])
        if n_it_1 == 0:
            n_it_1 = count
        elif n_it_2 == 0:
            n_it_2 = count

    return shop


def set_year_end_schedule(shop, target):
    starttime = shop.get_time_resolution()['timeresolution'].index[0]
    endtime = shop.get_time_resolution()['timeresolution'].index[-1]

    current_resolution = shop.get_time_resolution()['timeresolution'][0]

    date_range = pd.date_range(starttime, endtime, freq= str(current_resolution)+"H")

    schedule = []
    schedule_flag =[]

    for d in date_range:
        if d.month == 12 and d.day == 31 and d.hour == 24 - current_resolution:
            schedule.append(target)
            schedule_flag.append(2)
        else:
            schedule.append(0)
            schedule_flag.append(0)

    return schedule, schedule_flag, date_range
